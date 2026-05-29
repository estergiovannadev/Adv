from bs4 import BeautifulSoup
import re
import os
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import time

# ── Configuração ──────────────────────────────────────────────────────────────
OUTPUT       = r"Z:\Ester-Dev\advogados.xlsx"
DATA_INICIO  = "2024-01-01"
DATA_FIM     = date.today().isoformat()
URL_BASE     = (
    "https://comunica.pje.jus.br/consulta"
    "?dataDisponibilizacaoInicio={inicio}"
    "&dataDisponibilizacaoFim={fim}"
    "&numeroProcesso={numero}"
)

# OABs do escritório — esses advogados NUNCA entram na planilha
OABS_IGNORAR = {"MG-176171", "MG-124826"}
# ─────────────────────────────────────────────────────────────────────────────


def get_html_via_selenium(url):
    options = Options()
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1400,900")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })

    try:
        print(f"  Abrindo: {url}")
        driver.get(url)
        print("  Aguardando resultados carregarem...")
        for i in range(20):
            time.sleep(2)
            html = driver.page_source
            if "col-md-10" in html or "Advogado" in html:
                print(f"  Conteúdo detectado após {(i+1)*2}s.")
                break
            print(f"  Aguardando... {(i+1)*2}s")
        else:
            print("  AVISO: conteúdo não detectado no tempo limite.")

        html = driver.page_source
        with open(r"Z:\Ester-Dev\diagnostico.html", "w", encoding="utf-8") as f:
            f.write(html)
        return html
    finally:
        driver.quit()


def parse_dados(html, numero_processo):
    soup = BeautifulSoup(html, "html.parser")

    # ── 1. Pega o nome do Polo Ativo (autor) ─────────────────────────────────
    autor = ""
    for div in soup.find_all("div", class_="info-sumary"):
        polo_span = div.find("span", class_="tooltip-text")
        if polo_span and "Polo Ativo" in polo_span.text:
            spans = div.find_all("span")
            for i, s in enumerate(spans):
                if "Polo Ativo" in s.text and i + 1 < len(spans):
                    autor = spans[i + 1].get_text(strip=True)
                    break

    print(f"  Autor (Polo Ativo): {autor or 'não encontrado'}")

    # ── 2. Coleta todos os advogados listados ─────────────────────────────────
    todos = []
    vistos = set()
    for div in soup.find_all("div"):
        if "col-md-10" in div.get("class", []):
            raw = div.get_text(strip=True)
            match = re.match(r"^(.+?)\s*-\s*OAB\s+([A-Z]{2})-(\d+)$", raw)
            if match:
                oab_key = f"{match.group(2)}-{match.group(3)}"
                if oab_key not in vistos:
                    vistos.add(oab_key)
                    todos.append({
                        "nome": match.group(1).strip(),
                        "uf":   match.group(2).strip(),
                        "num":  match.group(3).strip(),
                        "oab":  oab_key,
                    })

    print(f"  Todos advogados: {[a['nome'] for a in todos]}")

    # ── 3. Remove OABs do escritório ─────────────────────────────────────────
    advogados_autor = [a for a in todos if a["oab"] not in OABS_IGNORAR]

    print(f"  Após filtro:     {[a['nome'] for a in advogados_autor]}")

    if not advogados_autor:
        print("  AVISO: nenhum advogado do autor encontrado após filtro.")
        return []

    # ── 4. Monta registros finais ─────────────────────────────────────────────
    return [{
        "Processo":    numero_processo,
        "Autor":       autor,
        "Advogado":    a["nome"],
        "UF OAB":      a["uf"],
        "Número OAB":  a["num"],
    } for a in advogados_autor]


def inicializar_xlsx(path):
    """Cria a planilha do zero com cabeçalho correto."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Advogados"

    headers = ["Processo", "Autor", "Advogado", "UF OAB", "Número OAB"]
    header_fill = PatternFill("solid", start_color="003366")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 15
    ws.row_dimensions[1].height = 20

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(path)
    print(f"  Nova planilha criada em: {path}")


def save_xlsx(registros, path):
    headers = ["Processo", "Autor", "Advogado", "UF OAB", "Número OAB"]
    row_font = Font(name="Arial", size=11)

    if not os.path.exists(path):
        inicializar_xlsx(path)

    wb = load_workbook(path)
    ws = wb.active
    next_row = ws.max_row + 1

    for reg in registros:
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=reg.get(h, ""))
            cell.font = row_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        next_row += 1

    wb.save(path)
    print(f"  Planilha salva em: {path}")


# ── Main ──────────────────────────────────────────────────────────────────────
print("=" * 50)
print("  Extrator de Advogados do Autor - PJe")
print("=" * 50)

numero_raw = input("\nDigite o número do processo: ").strip()
numero = re.sub(r"[\s.\-/]", "", numero_raw)

if not numero:
    print("Nenhum número informado. Encerrando.")
else:
    numero_fmt = (
        f"{numero[:7]}-{numero[7:9]}.{numero[9:13]}.{numero[13]}.{numero[14:16]}.{numero[16:]}"
        if len(numero) == 20 else numero_raw
    )
    print(f"\nNúmero formatado : {numero_fmt}")
    print(f"Período de busca : {DATA_INICIO} até {DATA_FIM}")

    url = URL_BASE.format(inicio=DATA_INICIO, fim=DATA_FIM, numero=numero)

    try:
        html = get_html_via_selenium(url)
        registros = parse_dados(html, numero_fmt)

        if registros:
            save_xlsx(registros, OUTPUT)
            print(f"\n✓ Adicionado(s) {len(registros)} advogado(s) do autor:")
            for r in registros:
                print(f"  Autor   : {r['Autor']}")
                print(f"  Advogado: {r['Advogado']} | OAB {r['UF OAB']}-{r['Número OAB']}")
        else:
            print("\nNenhum advogado do autor encontrado.")
            print("Verifique o diagnostico.html em Z:\\Ester-Dev\\")

    except Exception as e:
        print(f"\nErro: {e}")

input("\nPressione Enter para fechar...")
